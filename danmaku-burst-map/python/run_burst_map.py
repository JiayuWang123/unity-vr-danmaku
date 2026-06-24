#!/usr/bin/env python3
"""XML-only danmaku burst-map analyzer.

The tool accepts a Bilibili danmaku XML file and generates normalized data,
burst event tables, evidence-based topic/emotion/content labels, and report
assets. Optional plotting dependencies improve PNG output, but the core CSV,
JSON, Markdown, and HTML outputs use only the Python standard library.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import math
import os
import re
import statistics
import sys
import textwrap
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable, Sequence


SHORT_USEFUL_REACTIONS = {
    "牛逼",
    "进了",
    "好球",
    "漂亮",
    "绝了",
    "稳了",
    "燃",
    "帅",
    "强",
    "nb",
    "NB",
    "gg",
    "GG",
    "goal",
}

STOPWORDS = {
    "这个",
    "那个",
    "就是",
    "什么",
    "不是",
    "一个",
    "一下",
    "真的",
    "感觉",
    "还是",
    "可以",
    "没有",
    "这么",
    "怎么",
    "已经",
    "时候",
    "哈哈",
    "哈哈哈",
    "啊啊",
    "弹幕",
}

EMOTION_RULES = {
    "excitement": ["牛", "赢", "帅", "神", "漂亮", "nb", "NB", "冠军", "泪目", "哭了", "加油", "燃", "爽", "恭喜"],
    "tension": ["别", "稳住", "危险", "完了", "要寄", "别急", "吓", "紧张", "怕", "悬", "难了"],
    "mockery": ["笑死", "哈哈", "尴尬", "送", "菜", "回家", "搞笑", "小丑", "急了", "幽默"],
    "anger_or_dispute": ["裁判", "黑", "凭什么", "？？", "不是", "离谱", "恶心", "喷", "骂", "滚", "傻", "脑子"],
    "nostalgia_or_recap": ["去年", "今年", "最后", "到此一游", "离队", "回看", "纪录", "再见", "退役", "历史"],
    "viewer_behavior": ["去人声", "刷屏", "举报", "带节奏", "同一个人", "人声", "发弹幕", "打卡"],
    "neutral_analysis": ["阵容", "战术", "发育", "经济", "视野", "规则", "复盘", "选择", "版本"],
}

CONTENT_RULES = {
    "gameplay_reaction": ["一血", "击杀", "团", "开团", "拆", "基地", "推", "龙", "大龙", "小龙", "先锋", "进球", "扣篮", "绝杀"],
    "player_or_team_evaluation": ["队", "选手", "上单", "打野", "中单", "辅助", "Faker", "T1", "KT", "guma", "keria", "梅西", "詹姆斯"],
    "tactical_or_rule_discussion": ["BP", "bp", "阵容", "ban", "pick", "选", "战术", "规则", "越位", "判罚", "暂停", "换人"],
    "meme_or_slang": ["梗", "节目", "皮肤", "神", "享受", "大飞", "兰子", "抽象", "名场面"],
    "spam_or_chat_behavior": ["去人声", "刷屏", "举报", "弹幕", "带节奏", "同一个人", "人声", "发送"],
    "result_or_record_recap": ["打卡", "到此一游", "去年", "今年", "最后", "离队", "回看", "纪录", "冠军", "夺冠", "助攻", "再见"],
    "controversy_or_argument": ["裁判", "黑", "喷", "骂", "傻", "脑子", "恶心", "不是", "离谱", "怪", "节奏"],
}

BURST_KIND_RULES = {
    "viewer_behavior_peak": ["去人声", "刷屏", "举报", "同一个人", "人声"],
    "opening_artifact_peak": ["打卡", "到此一游", "来了", "第一", "考古", "补档"],
    "result_or_celebration_peak": ["冠军", "夺冠", "赢了", "捧杯", "合影", "纪录", "六冠", "绝杀"],
    "chat_meta_peak": ["弹幕", "带节奏", "举报", "刷屏"],
}


@dataclass
class DanmakuEntry:
    time_seconds: float
    mode: int
    font_size: int
    color: str
    timestamp: str
    pool: str
    user_hash: str
    danmaku_id: str
    raw_text: str
    clean_text: str
    normalized_text: str
    is_duplicate_like: bool
    is_spam_like: bool
    is_symbol_only: bool
    text_length: int
    priority_score: float
    evidence_weight: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate XML-only danmaku burst-map deliverables.")
    parser.add_argument("--input", required=True, help="Path to a Bilibili danmaku XML file.")
    parser.add_argument("--output", required=True, help="Output directory.")
    parser.add_argument("--config", default="", help="Optional config file path. Current parser uses built-in defaults.")
    parser.add_argument("--max-bursts", type=int, default=10, help="Maximum burst events to keep.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    entries = parse_bilibili_xml(input_path)
    if not entries:
        raise SystemExit(f"No valid danmaku entries parsed from {input_path}")

    density5 = build_density(entries, 5)
    density10 = build_density(entries, 10)
    burst_info = detect_bursts(density5, max_bursts=args.max_bursts)
    bursts = characterize_bursts(burst_info["bursts"], entries, burst_info["stats"])

    write_csv(output_dir / "normalized_danmaku.csv", [asdict(e) for e in entries])
    write_csv(output_dir / "density_5s.csv", density5)
    write_csv(output_dir / "density_10s.csv", density10)
    write_csv(output_dir / "burst_events.csv", [event_row(b) for b in bursts])
    write_csv(output_dir / "burst_characterization.csv", [characterization_row(b) for b in bursts])
    write_json(output_dir / "normalized_danmaku.json", [asdict(e) for e in entries])
    write_json(
        output_dir / "burst_events.json",
        {
            "source": {"input_xml": public_path_label(input_path), "window_seconds": 5, "xml_only": True},
            "stats": burst_info["stats"],
            "bursts": bursts,
        },
    )
    write_report(output_dir / "analysis_report.md", input_path, burst_info["stats"], bursts)
    write_html_table(output_dir / "burst_events_pretty.html", bursts)
    write_optional_xlsx(output_dir / "analysis_results.xlsx", entries, density5, bursts)
    make_charts(output_dir, density5, bursts, burst_info["stats"])

    print(f"Generated {len(bursts)} burst events from {len(entries)} XML danmaku entries.")
    print(f"Output: {output_dir}")
    return 0


def parse_bilibili_xml(path: Path) -> list[DanmakuEntry]:
    tree = ET.parse(path)
    raw_rows = []
    for node in tree.iter("d"):
        p = node.attrib.get("p", "")
        fields = p.split(",")
        if len(fields) < 8:
            continue
        try:
            time_seconds = float(fields[0])
            mode = int(float(fields[1]))
        except ValueError:
            continue
        raw_text = html.unescape(node.text or "")
        clean = clean_text(raw_text)
        normalized = normalize_for_duplicate(clean)
        raw_rows.append(
            {
                "time_seconds": time_seconds,
                "mode": mode,
                "font_size": safe_int(fields[2], 25),
                "color": color_to_hex(fields[3]),
                "timestamp": fields[4],
                "pool": fields[5],
                "user_hash": fields[6],
                "danmaku_id": fields[7],
                "raw_text": raw_text,
                "clean_text": clean,
                "normalized_text": normalized,
            }
        )

    duplicate_counts = Counter(row["normalized_text"] for row in raw_rows if row["normalized_text"])
    entries = []
    for row in raw_rows:
        text = row["clean_text"]
        symbol_only = bool(re.fullmatch(r"[\W_]+", text, flags=re.UNICODE)) if text else True
        dup_count = duplicate_counts.get(row["normalized_text"], 0)
        is_spam = dup_count >= 8 or repeated_pattern(text)
        priority = priority_score(text, row["mode"], row["font_size"], dup_count)
        entries.append(
            DanmakuEntry(
                **row,
                is_duplicate_like=dup_count > 1,
                is_spam_like=is_spam,
                is_symbol_only=symbol_only,
                text_length=len(text),
                priority_score=priority,
                evidence_weight=max(0.1, priority / (1.0 + math.log1p(max(0, dup_count - 1)))),
            )
        )
    entries.sort(key=lambda item: item.time_seconds)
    return entries


def public_path_label(path: Path) -> str:
    """Return a share-safe path label for reports and committed examples."""
    try:
        return path.relative_to(Path.cwd()).as_posix()
    except ValueError:
        return path.name


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(text or "")).strip()


def normalize_for_duplicate(text: str) -> str:
    return re.sub(r"\s+", "", text.lower())


def safe_int(value: str, default: int) -> int:
    try:
        return int(float(value))
    except ValueError:
        return default


def color_to_hex(value: str) -> str:
    try:
        return f"#{max(0, min(0xFFFFFF, int(float(value)))):06X}"
    except ValueError:
        return "#FFFFFF"


def repeated_pattern(text: str) -> bool:
    if not text:
        return False
    return bool(re.fullmatch(r"(.{1,4})\1{3,}", text))


def priority_score(text: str, mode: int, font_size: int, duplicate_count: int) -> float:
    if not text:
        return 0.0
    score = 1.0
    score += min(len(text) / 12.0, 2.0)
    if 4 <= len(text) <= 60:
        score += 1.0
    if any("\u4e00" <= ch <= "\u9fff" for ch in text):
        score += 0.35
    if any(cue in text.lower() for cue in ["哈", "笑", "牛", "帅", "好", "绝", "哭", "!", "?", "？"]):
        score += 0.45
    if mode in (4, 5):
        score += 0.25
    if font_size > 25:
        score += 0.15
    if duplicate_count >= 8:
        score -= 0.8
    if len(text) <= 1 and text not in SHORT_USEFUL_REACTIONS:
        score -= 1.0
    return max(0.0, round(score, 3))


def build_density(entries: Sequence[DanmakuEntry], window_seconds: int) -> list[dict]:
    max_time = max(e.time_seconds for e in entries)
    window_count = int(math.ceil(max_time / window_seconds)) + 1
    counts = [0] * window_count
    for entry in entries:
        idx = int(entry.time_seconds // window_seconds)
        counts[idx] += 1
    rows = []
    for idx, count in enumerate(counts):
        start = idx * window_seconds
        end = start + window_seconds
        rows.append(
            {
                "window_start_seconds": start,
                "window_end_seconds": end,
                "time_label": f"{seconds_label(start)}-{seconds_label(end)}",
                "raw_count": count,
                "count_per_second": round(count / window_seconds, 3),
            }
        )
    return rows


def detect_bursts(density: Sequence[dict], max_bursts: int) -> dict:
    counts = [int(row["raw_count"]) for row in density]
    starts = [float(row["window_start_seconds"]) for row in density]
    duration = float(density[-1]["window_end_seconds"])
    opening_guard = max(60.0, duration * 0.02)
    eligible = [count for count, start in zip(counts, starts) if start >= opening_guard] or counts
    median = statistics.median(eligible)
    mad = statistics.median([abs(x - median) for x in eligible])
    robust_sigma = 1.4826 * mad
    p90 = percentile(eligible, 90)
    p95 = percentile(eligible, 95)
    candidate_threshold = max(p90, median + 3 * robust_sigma)
    strong_threshold = max(p95, median + 6 * robust_sigma)
    hot = [idx for idx, count in enumerate(counts) if count >= candidate_threshold]
    groups = group_hot_windows(hot, density, merge_gap_seconds=10)
    bursts = []
    for group in groups:
        peak_idx = max(group, key=lambda idx: counts[idx])
        start = float(density[group[0]]["window_start_seconds"])
        end = float(density[group[-1]]["window_end_seconds"])
        peak = float(density[peak_idx]["window_start_seconds"])
        peak_count = counts[peak_idx]
        bursts.append(
            {
                "burst_id": "",
                "start_seconds": start,
                "end_seconds": end,
                "peak_seconds": peak,
                "time_range": f"{seconds_label(start)}-{seconds_label(end)}",
                "peak_density_5s": peak_count,
                "duration_seconds": round(end - start, 3),
                "total_raw_count": sum(counts[idx] for idx in group),
                "baseline_multiplier": round(peak_count / max(statistics.mean(eligible), 1e-6), 3),
                "is_opening_peak": start < opening_guard,
            }
        )

    def ranking_score(burst: dict) -> float:
        penalty = max([b["peak_density_5s"] for b in bursts], default=0) * 2 if burst["is_opening_peak"] else 0
        return burst["peak_density_5s"] - penalty

    bursts = sorted(bursts, key=ranking_score, reverse=True)[:max_bursts]
    bursts.sort(key=lambda item: item["start_seconds"])
    for idx, burst in enumerate(bursts, 1):
        burst["burst_id"] = f"B-{idx}"
    return {
        "stats": {
            "raw_total_count": sum(counts),
            "median_5s_count": median,
            "mad_5s_count": mad,
            "p90_5s_count": p90,
            "p95_5s_count": p95,
            "candidate_threshold": round(candidate_threshold, 3),
            "strong_threshold": round(strong_threshold, 3),
            "opening_guard_seconds": round(opening_guard, 3),
        },
        "bursts": bursts,
    }


def group_hot_windows(indices: Sequence[int], density: Sequence[dict], merge_gap_seconds: float) -> list[list[int]]:
    groups: list[list[int]] = []
    for idx in indices:
        if not groups:
            groups.append([idx])
            continue
        previous_end = float(density[groups[-1][-1]]["window_end_seconds"])
        current_start = float(density[idx]["window_start_seconds"])
        if current_start - previous_end <= merge_gap_seconds:
            groups[-1].append(idx)
        else:
            groups.append([idx])
    return groups


def characterize_bursts(bursts: Sequence[dict], entries: Sequence[DanmakuEntry], stats: dict) -> list[dict]:
    characterized = []
    for burst in bursts:
        context_start = max(0.0, burst["start_seconds"] - 10)
        context_end = burst["end_seconds"] + 15
        raw_context = [entry for entry in entries if context_start <= entry.time_seconds <= context_end]
        evidence_entries = [
            entry
            for entry in raw_context
            if entry.clean_text and not entry.is_symbol_only and entry.evidence_weight >= 0.5
        ]
        if not evidence_entries:
            evidence_entries = raw_context
        terms = top_terms([entry.clean_text for entry in evidence_entries], top_n=12)
        representatives = representative_comments(evidence_entries, terms, limit=5)
        emotion = classify_rules(evidence_entries, EMOTION_RULES)
        content = classify_content_mix(evidence_entries)
        topic_label, topic_confidence = infer_topic_label(terms, representatives, content, emotion)
        burst_kind = infer_burst_kind(burst, terms, representatives, content, stats["opening_guard_seconds"])
        characterized.append(
            {
                **burst,
                "burst_kind": burst_kind,
                "topic_label": topic_label,
                "topic_confidence": topic_confidence,
                "dominant_emotion": emotion["label"],
                "emotion_confidence": emotion["confidence_score"],
                "emotion_scores": emotion["label_scores"],
                "content_mix": content["mix"],
                "content_confidence": content["confidence_score"],
                "evidence_terms": "; ".join(f"{term}({count})" for term, count in terms[:8]),
                "evidence_comments": " | ".join(representatives[:3]),
                "representative_comments": " | ".join(representatives),
                "filtered_context_count": len(evidence_entries),
                "raw_context_count": len(raw_context),
            }
        )
    return characterized


def tokenize(text: str) -> list[str]:
    tokens: list[str] = []
    tokens.extend(re.findall(r"[A-Za-z][A-Za-z0-9_]{1,}", text))
    tokens.extend(re.findall(r"[\u4e00-\u9fff]{2,6}", text))
    for phrase in re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,12}", text):
        if len(phrase) >= 2:
            tokens.append(phrase)
    return [token for token in tokens if token not in STOPWORDS and len(token.strip()) >= 2]


def top_terms(texts: Iterable[str], top_n: int) -> list[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for text in texts:
        counter.update(tokenize(text))
    return counter.most_common(top_n)


def representative_comments(entries: Sequence[DanmakuEntry], terms: Sequence[tuple[str, int]], limit: int) -> list[str]:
    term_list = [term for term, _ in terms[:6]]
    scored = []
    for entry in entries:
        text = entry.clean_text
        if not text:
            continue
        term_bonus = sum(1 for term in term_list if term.lower() in text.lower())
        length_bonus = 1.0 if 4 <= len(text) <= 80 else 0.0
        score = entry.evidence_weight + term_bonus * 0.7 + length_bonus
        scored.append((score, text))
    output = []
    seen = set()
    for _, text in sorted(scored, reverse=True):
        key = normalize_for_duplicate(text)
        if key in seen:
            continue
        seen.add(key)
        output.append(text.replace("|", "/"))
        if len(output) >= limit:
            break
    return output


def classify_rules(entries: Sequence[DanmakuEntry], rules: dict[str, list[str]]) -> dict:
    counts: dict[str, int] = {}
    evidence_counts: Counter[str] = Counter()
    texts = [entry.clean_text for entry in entries]
    for label, terms in rules.items():
        label_count = 0
        for term in terms:
            hits = sum(1 for text in texts if term.lower() in text.lower())
            if hits:
                evidence_counts[term] += hits
                label_count += hits
        counts[label] = label_count
    total = sum(counts.values())
    if total <= 0:
        return {
            "label": "neutral_analysis",
            "confidence_score": 0.0,
            "evidence_counts": {},
            "label_counts": counts,
            "label_scores": {label: 0.0 for label in rules},
        }
    label = max(counts, key=counts.get)
    return {
        "label": label,
        "confidence_score": round(counts[label] / total, 3),
        "evidence_counts": dict(evidence_counts.most_common(8)),
        "label_counts": counts,
        "label_scores": {key: round(value / total, 3) for key, value in counts.items()},
    }


def classify_content_mix(entries: Sequence[DanmakuEntry]) -> dict:
    counts = {}
    texts = [entry.clean_text for entry in entries]
    for label, terms in CONTENT_RULES.items():
        counts[label] = sum(1 for text in texts for term in terms if term.lower() in text.lower())
    total = sum(counts.values())
    if total <= 0:
        return {"mix": "uncategorized_peak 100%", "confidence_score": 0.0, "primary": "uncategorized_peak"}
    parts = []
    for label, count in sorted(counts.items(), key=lambda item: item[1], reverse=True):
        if count > 0:
            parts.append(f"{label} {count / total:.0%}")
    primary = max(counts, key=counts.get)
    return {"mix": " / ".join(parts[:5]), "confidence_score": round(counts[primary] / total, 3), "primary": primary}


def infer_topic_label(terms: Sequence[tuple[str, int]], representatives: Sequence[str], content: dict, emotion: dict) -> tuple[str, float]:
    evidence = " ".join([term for term, _ in terms[:8]] + list(representatives[:2])).lower()
    if has_any(evidence, ["去人声", "人声", "刷屏", "举报"]):
        return "Viewer audio-removal spam and chat-order dispute", 0.9
    if has_any(evidence, ["冠军", "夺冠", "六冠", "捧杯", "合影", "纪录"]):
        return "Championship, record, and post-match celebration", 0.85
    if has_any(evidence, ["多兰", "兰子"]) and has_any(evidence, ["开团", "找机会", "送", "怪"]):
        return "Player decision recap and mockery around a key engage", 0.8
    if has_any(evidence, ["阵容", "bp", "ban", "pick", "选", "小火龙", "加里奥"]):
        return "Draft, tactics, and character/player choice discussion", 0.75
    if content["primary"] == "player_or_team_evaluation":
        return "Player and team performance evaluation", 0.7
    if content["primary"] == "gameplay_reaction":
        return "Gameplay event reaction", 0.65
    if terms:
        return "Focused discussion around " + ", ".join(term for term, _ in terms[:3]), 0.55
    return "Low-confidence topic", 0.1


def infer_burst_kind(burst: dict, terms: Sequence[tuple[str, int]], representatives: Sequence[str], content: dict, opening_guard: float) -> str:
    evidence = " ".join([term for term, _ in terms[:8]] + list(representatives[:3])).lower()
    if burst["start_seconds"] < opening_guard and has_any(evidence, BURST_KIND_RULES["opening_artifact_peak"]):
        return "opening_artifact_peak"
    if has_any(evidence, BURST_KIND_RULES["viewer_behavior_peak"]):
        return "viewer_behavior_peak"
    if has_any(evidence, BURST_KIND_RULES["result_or_celebration_peak"]):
        return "result_or_celebration_peak"
    if content["primary"] in ("gameplay_reaction", "tactical_or_rule_discussion", "player_or_team_evaluation"):
        return "gameplay_peak"
    if has_any(evidence, BURST_KIND_RULES["chat_meta_peak"]):
        return "chat_meta_peak"
    return "uncategorized_peak"


def has_any(text: str, terms: Sequence[str]) -> bool:
    return any(term.lower() in text for term in terms)


def event_row(burst: dict) -> dict:
    keys = [
        "burst_id",
        "time_range",
        "start_seconds",
        "end_seconds",
        "peak_seconds",
        "peak_density_5s",
        "duration_seconds",
        "baseline_multiplier",
        "burst_kind",
        "topic_label",
        "dominant_emotion",
        "content_mix",
        "filtered_context_count",
        "raw_context_count",
    ]
    return {key: burst.get(key, "") for key in keys}


def characterization_row(burst: dict) -> dict:
    keys = [
        "burst_id",
        "time_range",
        "topic_label",
        "topic_confidence",
        "dominant_emotion",
        "emotion_confidence",
        "content_mix",
        "content_confidence",
        "evidence_terms",
        "evidence_comments",
        "representative_comments",
    ]
    row = {key: burst.get(key, "") for key in keys}
    row["representative_comments_short"] = shorten(row["representative_comments"], 120)
    return row


def write_csv(path: Path, rows: Sequence[dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_report(path: Path, input_path: Path, stats: dict, bursts: Sequence[dict]) -> None:
    lines = [
        "# Danmaku Burst Map Analysis Report",
        "",
        f"Input XML: `{input_path.name}`",
        "",
        "## Density Baseline",
        "",
        f"- Raw danmaku count: {stats['raw_total_count']}",
        f"- Candidate threshold: {stats['candidate_threshold']} comments / 5s",
        f"- Strong threshold: {stats['strong_threshold']} comments / 5s",
        f"- Opening guard: {stats['opening_guard_seconds']} seconds",
        "",
        "## Burst Events",
        "",
        "| ID | Time Range | Peak / 5s | Duration | Kind | Topic | Emotion |",
        "|---|---|---:|---:|---|---|---|",
    ]
    for burst in bursts:
        lines.append(
            f"| {burst['burst_id']} | {burst['time_range']} | {burst['peak_density_5s']} | "
            f"{burst['duration_seconds']} | {burst['burst_kind']} | {burst['topic_label']} | {burst['dominant_emotion']} |"
        )
    lines.extend(["", "## Evidence Notes", ""])
    for burst in bursts:
        lines.extend(
            [
                f"### {burst['burst_id']} {burst['time_range']}",
                "",
                f"- Topic evidence terms: {burst['evidence_terms']}",
                f"- Representative comments: {burst['representative_comments']}",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_html_table(path: Path, bursts: Sequence[dict]) -> None:
    rows = "\n".join(
        "<tr>"
        + "".join(
            f"<td>{html.escape(str(burst.get(key, '')))}</td>"
            for key in ["burst_id", "time_range", "peak_density_5s", "burst_kind", "topic_label", "dominant_emotion"]
        )
        + "</tr>"
        for burst in bursts
    )
    path.write_text(
        f"""<!doctype html>
<html><head><meta charset="utf-8"><title>Danmaku Burst Events</title>
<style>body{{font-family:Arial,'Microsoft YaHei','SimHei','DengXian','Noto Sans CJK SC',sans-serif;margin:24px}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ddd;padding:8px;vertical-align:top}}th{{background:#f3f5f7;position:sticky;top:0}}tr:nth-child(even){{background:#fafafa}}</style>
</head><body><h1>Danmaku Burst Events</h1><table>
<thead><tr><th>ID</th><th>Time Range</th><th>Peak / 5s</th><th>Kind</th><th>Topic</th><th>Emotion</th></tr></thead>
<tbody>{rows}</tbody></table></body></html>""",
        encoding="utf-8",
    )


def write_optional_xlsx(path: Path, entries: Sequence[DanmakuEntry], density: Sequence[dict], bursts: Sequence[dict]) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        return
    workbook = openpyxl.Workbook()
    sheets = [
        ("Normalized Danmaku", [asdict(e) for e in entries[:5000]]),
        ("Density", list(density)),
        ("Burst Events", [event_row(b) for b in bursts]),
        ("Characterization", [characterization_row(b) for b in bursts]),
    ]
    workbook.remove(workbook.active)
    fills = {
        "viewer_behavior_peak": "FFF2CC",
        "result_or_celebration_peak": "D9EAD3",
        "gameplay_peak": "D9EAF7",
        "opening_artifact_peak": "EADCF8",
    }
    for name, rows in sheets:
        ws = workbook.create_sheet(name)
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F3F5F7")
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.font = Font(name="Microsoft YaHei")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if "burst_kind" in headers:
            kind_col = headers.index("burst_kind") + 1
            for row_idx in range(2, ws.max_row + 1):
                kind = ws.cell(row_idx, kind_col).value
                if kind in fills:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=fills[kind])
        for column in ws.columns:
            max_len = min(48, max(len(str(cell.value or "")) for cell in column) + 2)
            ws.column_dimensions[column[0].column_letter].width = max(12, max_len)
    workbook.save(path)


def make_charts(output_dir: Path, density: Sequence[dict], bursts: Sequence[dict], stats: dict) -> None:
    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        write_svg_fallback(output_dir / "density_curve_5s.svg", density, bursts, stats)
        return
    configure_matplotlib_fonts(plt)
    times = [row["window_start_seconds"] / 60 for row in density]
    counts = [row["raw_count"] for row in density]
    plt.figure(figsize=(12, 5))
    plt.plot(times, counts, linewidth=1.2)
    plt.axvspan(0, stats["opening_guard_seconds"] / 60, color="lightgray", alpha=0.35, label="Opening guard")
    plt.axhline(stats["candidate_threshold"], linestyle="--", color="#E67E22", label="Candidate threshold")
    plt.axhline(stats["strong_threshold"], linestyle="--", color="#C0392B", label="Strong threshold")
    for burst in bursts:
        x = burst["peak_seconds"] / 60
        y = burst["peak_density_5s"]
        plt.scatter([x], [y], s=42)
        plt.text(x, y + 1, burst["burst_id"], ha="center", fontsize=9)
    plt.title("Danmaku Density Over Time (5s Window)")
    plt.xlabel("Video Time (minutes)")
    plt.ylabel("Danmaku Count / 5s")
    plt.grid(alpha=0.25)
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / "density_curve_5s.png", dpi=180)
    plt.close()

    bar_chart(output_dir / "burst_peak_bar_chart.png", bursts, "peak_density_5s", "Peak Density by Burst Event", "Danmaku Count / 5s", plt)
    bar_chart(output_dir / "burst_duration_chart.png", bursts, "duration_seconds", "Burst Duration by Event", "Duration (seconds)", plt)
    timeline_chart(output_dir / "burst_timeline.png", bursts, plt)
    summary_chart(output_dir / "emotion_content_summary.png", bursts, plt)
    heatmap_chart(output_dir / "burst_heatmap.png", bursts, plt)
    keyword_chart(output_dir / "topic_keyword_evidence_chart.png", bursts, plt)


def bar_chart(path: Path, bursts: Sequence[dict], metric: str, title: str, ylabel: str, plt) -> None:
    ordered = sorted(bursts, key=lambda item: item[metric], reverse=True)
    plt.figure(figsize=(10, 5))
    plt.bar([b["burst_id"] for b in ordered], [b[metric] for b in ordered])
    plt.title(title)
    plt.xlabel("Burst Event")
    plt.ylabel(ylabel)
    plt.grid(axis="y", alpha=0.25)
    plt.tight_layout()
    plt.savefig(path, dpi=180)
    plt.close()


def timeline_chart(path: Path, bursts: Sequence[dict], plt) -> None:
    colors = defaultdict(lambda: "#777777", {
        "gameplay_peak": "#4C78A8",
        "result_or_celebration_peak": "#54A24B",
        "viewer_behavior_peak": "#F58518",
        "opening_artifact_peak": "#B279A2",
        "chat_meta_peak": "#E45756",
    })
    friendly_kind = {
        "gameplay_peak": "Gameplay",
        "result_or_celebration_peak": "Result / celebration",
        "viewer_behavior_peak": "Viewer behavior",
        "opening_artifact_peak": "Opening artifact",
        "chat_meta_peak": "Chat meta",
        "uncategorized_peak": "Uncategorized",
    }
    ordered = sorted(bursts, key=lambda item: item["start_seconds"])
    if not ordered:
        return

    max_peak = max(float(b["peak_density_5s"]) for b in ordered) or 1.0
    x_min = max(0.0, min(float(b["start_seconds"]) for b in ordered) / 60 - 0.8)
    x_max = max(float(b["end_seconds"]) for b in ordered) / 60 + 0.8
    lanes = [1.35, -1.35, 2.35, -2.35]

    fig, ax = plt.subplots(figsize=(15, 6.6), facecolor="white")
    ax.axhline(0, color="#2F3542", linewidth=2.2, alpha=0.85)
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(-3.25, 3.25)
    ax.set_yticks([])
    ax.grid(axis="x", color="#D8DEE9", linewidth=0.9, alpha=0.7)
    ax.set_axisbelow(True)

    used_kinds = set()
    for idx, burst in enumerate(ordered):
        kind = burst["burst_kind"]
        color = colors[kind]
        used_kinds.add(kind)
        start = float(burst["start_seconds"]) / 60
        end = float(burst["end_seconds"]) / 60
        peak = float(burst["peak_seconds"]) / 60
        peak_density = float(burst["peak_density_5s"])
        duration = max(end - start, 0.035)
        lane = lanes[idx % len(lanes)]
        line_width = 5 + 10 * math.sqrt(peak_density / max_peak)
        marker_size = 90 + 520 * math.sqrt(peak_density / max_peak)

        ax.plot(
            [start, max(end, start + duration)],
            [0, 0],
            color=color,
            linewidth=line_width,
            solid_capstyle="round",
            alpha=0.50,
            zorder=2,
        )
        ax.scatter(
            [peak],
            [0],
            s=marker_size,
            color=color,
            edgecolors="white",
            linewidths=1.8,
            zorder=4,
        )
        ax.text(
            peak,
            0,
            burst["burst_id"],
            color="white",
            fontsize=9,
            fontweight="bold",
            ha="center",
            va="center",
            zorder=5,
        )

        topic = shorten(str(burst.get("topic_label", "")), 36)
        label = (
            f"{burst['burst_id']}  {burst['time_range']}\n"
            f"Peak {burst['peak_density_5s']}/5s · {friendly_kind.get(kind, kind)}\n"
            f"{topic}"
        )
        ax.annotate(
            label,
            xy=(peak, 0),
            xytext=(peak, lane),
            ha="center",
            va="center",
            fontsize=9,
            linespacing=1.15,
            bbox=dict(boxstyle="round,pad=0.42,rounding_size=0.15", fc="white", ec=color, lw=1.6, alpha=0.96),
            arrowprops=dict(arrowstyle="-", color=color, lw=1.3, alpha=0.85, shrinkA=4, shrinkB=8),
            zorder=6,
        )

    legend_handles = [
        plt.Line2D([0], [0], color=colors[kind], marker="o", linestyle="", markersize=9, label=friendly_kind.get(kind, kind))
        for kind in sorted(used_kinds)
    ]
    ax.legend(handles=legend_handles, loc="upper left", frameon=False, ncol=min(3, len(legend_handles)))
    ax.set_title("Burst Event Timeline", fontsize=16, fontweight="bold", pad=14)
    ax.set_xlabel("Video Time (minutes)")
    ax.text(
        x_min,
        -3.05,
        "Circle size encodes peak density; colored line length encodes burst duration.",
        fontsize=9,
        color="#57606F",
        ha="left",
    )
    for spine in ["left", "right", "top"]:
        ax.spines[spine].set_visible(False)
    ax.spines["bottom"].set_color("#A4B0BE")
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    plt.close()


def summary_chart(path: Path, bursts: Sequence[dict], plt) -> None:
    emotion_counts = Counter(b["dominant_emotion"] for b in bursts)
    content_counts = Counter((b["content_mix"].split(" ")[0] if b["content_mix"] else "uncategorized") for b in bursts)
    plt.figure(figsize=(12, 5))
    plt.subplot(1, 2, 1)
    plt.bar(emotion_counts.keys(), emotion_counts.values())
    plt.title("Dominant Emotion Summary")
    plt.xticks(rotation=35, ha="right")
    plt.subplot(1, 2, 2)
    plt.bar(content_counts.keys(), content_counts.values())
    plt.title("Primary Content Summary")
    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    plt.savefig(path, dpi=180)
    plt.close()


def heatmap_chart(path: Path, bursts: Sequence[dict], plt) -> None:
    labels = [label for label in EMOTION_RULES.keys() if any(b.get("emotion_scores", {}).get(label, 0) > 0 for b in bursts)]
    if not labels:
        labels = sorted(set(b["dominant_emotion"] for b in bursts))
    matrix = [
        [float(b.get("emotion_scores", {}).get(label, 1.0 if b["dominant_emotion"] == label else 0.0)) for b in bursts]
        for label in labels
    ]
    fig, ax = plt.subplots(figsize=(13, max(4, len(labels) * 0.68)), facecolor="white")
    image = ax.imshow(matrix, aspect="auto", cmap="YlOrRd", vmin=0, vmax=1)
    ax.set_xticks(range(len(bursts)))
    ax.set_xticklabels([b["burst_id"] for b in bursts])
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.set_title("Emotion Evidence Strength by Burst", fontsize=15, fontweight="bold", pad=12)
    ax.set_xlabel("Burst Event")
    ax.set_ylabel("Emotion Category")
    ax.set_xticks([x - 0.5 for x in range(1, len(bursts))], minor=True)
    ax.set_yticks([y - 0.5 for y in range(1, len(labels))], minor=True)
    ax.grid(which="minor", color="white", linewidth=1.2)
    ax.tick_params(which="minor", bottom=False, left=False)

    for y, row in enumerate(matrix):
        for x, value in enumerate(row):
            if value <= 0:
                continue
            color = "white" if value >= 0.45 else "#2F3542"
            ax.text(x, y, f"{value:.2f}", ha="center", va="center", fontsize=8, color=color, fontweight="bold" if value >= 0.45 else "normal")

    cbar = fig.colorbar(image, ax=ax, fraction=0.035, pad=0.025)
    cbar.set_label("Evidence strength (share of matched emotion cues)")
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)


def keyword_chart(path: Path, bursts: Sequence[dict], plt) -> None:
    top = Counter()
    for burst in bursts:
        for token in burst["evidence_terms"].split("; "):
            if "(" in token:
                term = token.split("(")[0]
                count = int(token.rsplit("(", 1)[1].rstrip(")"))
                top[term] += count
    common = top.most_common(12)
    plt.figure(figsize=(10, 6))
    plt.barh([term for term, _ in reversed(common)], [count for _, count in reversed(common)])
    plt.title("Keyword Evidence Across Burst Events")
    plt.xlabel("Evidence Count")
    plt.tight_layout()
    plt.savefig(path, dpi=180)
    plt.close()


def configure_matplotlib_fonts(plt) -> None:
    candidates = [
        "Microsoft YaHei",
        "SimHei",
        "DengXian",
        "Noto Sans CJK SC",
        "Noto Sans SC",
        "Arial Unicode MS",
    ]
    try:
        from matplotlib import font_manager  # type: ignore

        for font_path in [
            r"C:\Windows\Fonts\msyh.ttc",
            r"C:\Windows\Fonts\simhei.ttf",
            r"C:\Windows\Fonts\Deng.ttf",
            r"C:\Windows\Fonts\simsun.ttc",
        ]:
            if os.path.exists(font_path):
                font_manager.fontManager.addfont(font_path)
        installed = {font.name for font in font_manager.fontManager.ttflist}
        chosen = next((font for font in candidates if font in installed), None)
        if chosen:
            plt.rcParams["font.sans-serif"] = [chosen] + candidates
            plt.rcParams["font.family"] = "sans-serif"
    except Exception:
        plt.rcParams["font.sans-serif"] = candidates
    plt.rcParams["axes.unicode_minus"] = False


def write_svg_fallback(path: Path, density: Sequence[dict], bursts: Sequence[dict], stats: dict) -> None:
    width, height = 1200, 480
    counts = [row["raw_count"] for row in density]
    max_count = max(max(counts), 1)
    points = []
    for idx, count in enumerate(counts):
        x = 60 + idx / max(1, len(counts) - 1) * (width - 100)
        y = height - 50 - count / max_count * (height - 100)
        points.append(f"{x:.1f},{y:.1f}")
    path.write_text(
        f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}">
<rect width="100%" height="100%" fill="white"/>
<text x="60" y="32" font-family="Arial, Microsoft YaHei, SimHei, DengXian, sans-serif" font-size="22">Danmaku Density Over Time (5s Window)</text>
<polyline fill="none" stroke="#2f5f9f" stroke-width="2" points="{' '.join(points)}"/>
<text x="60" y="{height-18}" font-family="Arial, Microsoft YaHei, SimHei, DengXian, sans-serif" font-size="14">PNG charts require matplotlib; SVG fallback generated.</text>
</svg>""",
        encoding="utf-8",
    )


def seconds_label(seconds: float) -> str:
    seconds = max(0, float(seconds))
    return f"{int(seconds // 60):02d}:{int(seconds % 60):02d}"


def percentile(values: Sequence[float], pct: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    pos = pct / 100 * (len(ordered) - 1)
    low = math.floor(pos)
    high = math.ceil(pos)
    if low == high:
        return float(ordered[low])
    return float(ordered[low] + (ordered[high] - ordered[low]) * (pos - low))


def shorten(text: str, limit: int) -> str:
    text = text or ""
    return text if len(text) <= limit else text[: limit - 1] + "…"


if __name__ == "__main__":
    raise SystemExit(main())
